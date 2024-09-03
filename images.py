from openpyxl import load_workbook
from openai import OpenAI
from upload import main
import pandas as pd
import requests
import dotenv
import time
import math
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from PIL import Image
import io

dotenv.load_dotenv('.env')

# Initialize OpenAI client
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

def create_slug(keyword: str) -> str:
    return keyword.replace(" ", "-").lower()

def process_image(row, index, ws, field1_name, field2_name, is_generated_field_name, model, size, quality, wait_time=15):
    keyword = row[field1_name]
    design_description = row[field2_name]
    slug = create_slug(keyword)

    # Refined prompt to ensure a white background and no extra elements
    prompt = f"{keyword}, {design_description}"
    print("Prompt: ", prompt)

    attempts = 0
    while attempts < 3:  # Retry up to 3 times
        try:
            # Generate the image using OpenAI's DALL-E model
            response = client.images.generate(
                model=model,
                prompt=prompt,
                n=1,
                size=size,
            )

            image_url = response.data[0].url

            # Download the image data
            image_response = requests.get(image_url)
            image_data = image_response.content
            
            # Convert the image to .webp format and resize to 512x512
            image = Image.open(io.BytesIO(image_data))
            resized_image = image.resize((512, 512))  # Resize to 512x512
            webp_image_data = io.BytesIO()
            resized_image.save(webp_image_data, format='webp')
            webp_image_data.seek(0)

            # Save the image to the images folder
            image_filename = (f"{keyword}.webp")
            main(webp_image_data.read(), image_filename, slug)

            # Mark the row as generated
            ws[f'D{index + 2}'] = True
            return  # Exit the function on success

        except Exception as e:
            error_message = str(e)
            if 'rate_limit_exceeded' in error_message:
                print(f"Rate limit exceeded. Waiting for {wait_time} seconds before retrying...")
                time.sleep(wait_time)
                attempts += 1
            else:
                print(f"Failed to process image for keyword '{keyword}'. Error: {e}")
                break  # Non-retriable error, break out of the loop

def extract_prompt_field_data_and_download_images(excel_file, field1_name, field2_name, is_generated_field_name, model=os.getenv('DALLE'), size=os.getenv('DALLE_RESOLUTION'), quality="standard", max_requests_per_minute=15):
    """
    Extracts complete prompt field data by joining values from two other fields, generates images using the DALL-E model,
    uploads the images to the website, and updates the Excel file.
    It also checks the "isGenerated" field and skips image generation if it's already marked as true.

    Parameters:
    - excel_file: str, path to the Excel file
    - field1_name: str, name of the first field (used for generating filenames)
    - field2_name: str, name of the second field
    - is_generated_field_name: str, name of the field to check for image generation status
    - model: str, name of the DALL-E model to use
    - size: str, size of the generated image
    - quality: str, quality of the generated image
    - max_requests_per_minute: int, maximum number of requests per minute

    Returns:
    - None
    """
    # Read the Excel file
    df = pd.read_excel(excel_file)
    wb = load_workbook(excel_file)
    ws = wb.active

    requests_count = 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=15) as executor:
        futures = []
        for index, row in df.iterrows():
            if not row[is_generated_field_name]:
                futures.append(executor.submit(process_image, row, index, ws, field1_name, field2_name, is_generated_field_name, model, size, quality))

                requests_count += 1
                if requests_count >= max_requests_per_minute:
                    elapsed_time = time.time() - start_time
                    if elapsed_time < 60:
                        time_to_wait = 60 - elapsed_time
                        print(f"Reached {max_requests_per_minute} requests. Waiting for {time_to_wait} seconds...")
                        time.sleep(time_to_wait)
                    requests_count = 0
                    start_time = time.time()

        for future in as_completed(futures):
            future.result()
            # Save the workbook after each completed image processing
            wb.save(excel_file)
            print("Incrementally updated workbook saved.")

    # Final save of the workbook
    wb.save(excel_file)
    print("Final updated workbook saved.")

st = time.time()
# Example usage:
excel_file = "Images.xlsx"
field2_name = "Design Instructions"
field1_name = "Keyword"
is_generated_field_name = "Generated"
extract_prompt_field_data_and_download_images(excel_file, field1_name, field2_name, is_generated_field_name)

end = time.time()

print("Image generation and upload completed", math.floor((end-st)*100)/100," seconds")
